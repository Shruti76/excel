#!/usr/bin/env python3
"""
April Site Visit Calendar Analysis Script
Processes 5 types of input files and generates comprehensive April report

File Mapping:
1. Colo_Sonatel-APS.xlsx:
   - Columns: nearest Helios site ID, Nearest Helios site distance, Tower type, Site ID (YAS SID), Site Name, Projet
   - Used for: Base site reference mapping (Helios → YAS)

2. HTS Site Visits Calendar (Planning sheet):
   - Row 3-6 headers: Activity, site ID, Name, Location Type, Latitude, Longitude, (empty), (empty), Region, Mois, Technical Staff, Non-Tech 1, Non-Tech 2, Date, SARID
   - March data used as reference for April planning
   - Used for: Activity type, staff assignments, location details

3. PM_ZONE1_2_MARS_2026.xlsx:
   - Columns: YAS SID, HTS SID (Helios), Assigned to, Planning Date
   - Mars (March) dates - adjust to April
   - Used for: Preventive Maintenance visits with Lat/Lon from site master

4. Project sites BNS plan:
   - Columns include: Helios Site ID (SNKL...), YAS Site ID, Region, Activity, Lat, Long
   - Used for: Project-based site visits

5. RT High Risk List:
   - Identifies sites to EXCLUDE (dangerous/failed)
   - Used for: Filtering out risky sites from all sources
"""

import pandas as pd
import numpy as np
import sys
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class AprilReportGenerator:
    """Generate comprehensive April site visit schedule report from 5 file types."""
    
    def __init__(self, output_folder: str = "output"):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        self.log = []
        self.report_date = datetime(2026, 4, 1)  # April 2026
        
        # High-risk site IDs to exclude (from RT High Risk List)
        self.high_risk_site_ids = set()
        
        # Loaded data
        self.site_master = None
        self.hts_calendar = None
        self.pm_assignments = None
        self.project_sites = None
        self.critical_sites = None
        
        # Analysis results
        self.april_schedule = None
        self.team_assignments = None
        self.summary_stats = None
    
    def log_message(self, msg: str):
        """Log message with timestamp."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        full_msg = f"[{timestamp}] {msg}"
        self.log.append(full_msg)
        print(full_msg)
    
    def load_files(self, file_dict: Dict[str, str]) -> bool:
        """
        Load all 5 file types and extract high-risk site IDs to exclude.
        
        Args:
            file_dict: Dictionary with keys:
                - 'site_master': Colo_Sonatel-APS.xlsx (Helios Site ID → YAS SID)
                - 'hts_calendar': HTS Site Visits Calendar (Preventive Maintenance activity data)
                - 'pm_assignments': PM_ZONE1_2_MARS_2026.xlsx (Helios Site ID → YAS SID + Lat/Long)
                - 'project_sites': Project sites BNS plan (Helios Site ID → YAS ID + Lat/Long)
                - 'critical_sites': RT High Risk List (Sites to EXCLUDE - Helios/YAS IDs)
        """
        try:
            # Load Critical Sites FIRST (File 5) - to extract high-risk site IDs for exclusion
            self.high_risk_site_ids = set()
            if 'critical_sites' in file_dict:
                self.critical_sites = pd.read_excel(file_dict['critical_sites'], sheet_name=0)
                # Extract all Helios Site IDs and YAS IDs from high-risk list
                id_cols = [c for c in self.critical_sites.columns if 'id' in c.lower() or 'sid' in c.lower()]
                for col in id_cols:
                    self.high_risk_site_ids.update(
                        self.critical_sites[col].dropna().astype(str).unique()
                    )
                self.log_message(f"⚠️  Loaded High Risk Sites: {len(self.high_risk_site_ids)} sites to EXCLUDE")
            
            # Load site master (File 1)
            if 'site_master' in file_dict:
                self.site_master = pd.read_excel(file_dict['site_master'], sheet_name=0)
                # Filter out high-risk sites
                id_cols = [c for c in self.site_master.columns if 'id' in c.lower() or 'sid' in c.lower()]
                initial_count = len(self.site_master)
                for col in id_cols:
                    self.site_master = self.site_master[~self.site_master[col].astype(str).isin(self.high_risk_site_ids)]
                self.log_message(f"✅ Loaded Site Master: {len(self.site_master)} sites (excluded {initial_count - len(self.site_master)} high-risk)")
            
            # Load HTS Calendar (File 2) - Preventive Maintenance planning schedule
            if 'hts_calendar' in file_dict:
                self.hts_calendar = pd.read_excel(file_dict['hts_calendar'], sheet_name='Planning')
                # Filter out high-risk sites from HTS Calendar
                id_cols = [c for c in self.hts_calendar.columns if 'id' in c.lower() or 'sid' in c.lower()]
                initial_count = len(self.hts_calendar)
                for col in id_cols:
                    self.hts_calendar = self.hts_calendar[~self.hts_calendar[col].astype(str).isin(self.high_risk_site_ids)]
                self.log_message(f"✅ Loaded HTS Calendar: {len(self.hts_calendar)} records (excluded {initial_count - len(self.hts_calendar)} high-risk)")
            
            # Load PM Assignments (File 3) - Preventive Maintenance activity assignments
            if 'pm_assignments' in file_dict:
                self.pm_assignments = pd.read_excel(file_dict['pm_assignments'], sheet_name=0)
                # Filter out high-risk sites from PM assignments
                id_cols = [c for c in self.pm_assignments.columns if 'id' in c.lower() or 'sid' in c.lower()]
                initial_count = len(self.pm_assignments)
                for col in id_cols:
                    self.pm_assignments = self.pm_assignments[~self.pm_assignments[col].astype(str).isin(self.high_risk_site_ids)]
                self.log_message(f"✅ Loaded PM Assignments: {len(self.pm_assignments)} records (excluded {initial_count - len(self.pm_assignments)} high-risk)")
            
            # Load Project Sites (File 4)
            if 'project_sites' in file_dict:
                df = pd.read_excel(file_dict['project_sites'], sheet_name=0)
                # Clean up the dataframe - remove empty rows/columns
                df = df.dropna(how='all').dropna(axis=1, how='all')
                # Filter out high-risk sites from project sites
                id_cols = [c for c in df.columns if 'id' in c.lower() or 'sid' in c.lower()]
                initial_count = len(df)
                for col in id_cols:
                    df = df[~df[col].astype(str).isin(self.high_risk_site_ids)]
                self.project_sites = df
                self.log_message(f"✅ Loaded Project Sites: {len(self.project_sites)} sites (excluded {initial_count - len(df)} high-risk)")
            
            return True
        
        except Exception as e:
            self.log_message(f"❌ Error loading files: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def extract_site_ids(self) -> set:
        """Extract all unique site IDs from all sources."""
        site_ids = set()
        
        if self.site_master is not None:
            cols = [c for c in self.site_master.columns if 'id' in c.lower() or 'sid' in c.lower()]
            for col in cols:
                site_ids.update(self.site_master[col].dropna().unique())
        
        if self.pm_assignments is not None:
            cols = [c for c in self.pm_assignments.columns if 'id' in c.lower() or 'sid' in c.lower()]
            for col in cols:
                site_ids.update(self.pm_assignments[col].dropna().unique())
        
        if self.critical_sites is not None:
            cols = [c for c in self.critical_sites.columns if 'id' in c.lower() or 'sid' in c.lower()]
            for col in cols:
                site_ids.update(self.critical_sites[col].dropna().unique())
        
        self.log_message(f"📊 Total unique sites identified: {len(site_ids)}")
        return site_ids
    
    def _get_site_master_lookup(self) -> dict:
        """Create lookup dict for site master: YAS SID -> (Helios ID, Name, Tower Type, Project)"""
        lookup = {}
        if self.site_master is not None and len(self.site_master) > 0:
            for idx, row in self.site_master.iterrows():
                yas_sid = str(row.iloc[3]) if pd.notna(row.iloc[3]) else None  # Site ID (col 4)
                if yas_sid:
                    lookup[yas_sid] = {
                        'helios_id': row.iloc[0] if pd.notna(row.iloc[0]) else '',  # nearest Helios site ID
                        'name': row.iloc[4] if pd.notna(row.iloc[4]) else '',  # Site Name (col 5)
                        'tower_type': row.iloc[2] if pd.notna(row.iloc[2]) else '',  # Tower type (col 3)
                        'project': row.iloc[5] if pd.notna(row.iloc[5]) else '',  # Project (col 6)
                        'latitude': '',
                        'longitude': ''
                    }
        return lookup
    
    def _get_engineer_pool(self) -> dict:
        """Get engineer pools from March reference or define default pools."""
        # Tech staff pool
        tech_staff = [
            'Bara Diokhané', 'Boubacar Mbeguere', 'Ousman Diouf', 'Abdoul Aziz Seck'
        ]
        # Non-tech staff pool (Group 1 & 2)
        non_tech_pool = [
            'Koudjo Guy Gameda', 'Djienaba Fall', 'Victor Bocande', 'Cheikh Niang',
            'Mahib Samb', 'Marième Gno Dia', 'Hamedine Fall Thiam', 'Ahmad Toure'
        ]
        return {'tech': tech_staff, 'non_tech': non_tech_pool}
    
    def _assign_staff(self, visit_index: int, engineer_pools: dict) -> tuple:
        """Assign tech and non-tech staff based on visit index (rotates through pools)."""
        tech_staff = engineer_pools['tech']
        non_tech_pool = engineer_pools['non_tech']
        
        # Rotate through tech staff
        tech_idx = visit_index % len(tech_staff)
        tech = tech_staff[tech_idx]
        
        # Assign 2 non-tech staff (different pair for each tech staff)
        non_tech_start = (visit_index * 2) % len(non_tech_pool)
        non_tech_1 = non_tech_pool[non_tech_start]
        non_tech_2 = non_tech_pool[(non_tech_start + 1) % len(non_tech_pool)]
        
        return tech, non_tech_1, non_tech_2
    
    def generate_april_schedule(self) -> pd.DataFrame:
        """
        Generate April visit schedule combining all 5 sources with proper:
        - Activity types from each source
        - Engineer assignment for April only (not repeating throughout year)
        - 2 non-technical staff per visit like March reference
        - Varied dates throughout April
        """
        self.log_message("Generating comprehensive April schedule (all sources, April-only assignments)...")
        
        schedule_data = []
        processed_sites = set()
        site_master_lookup = self._get_site_master_lookup()
        engineer_pools = self._get_engineer_pool()
        visit_count = 0
        
        # ===== PRIORITY 1: PM_ZONE1_2 (Preventive Maintenance - from assigned_to column) =====
        if self.pm_assignments is not None and len(self.pm_assignments) > 0:
            self.log_message(f"  Processing {len(self.pm_assignments)} PM Assignments (Preventive Maintenance)...")
            for idx, row in self.pm_assignments.iterrows():
                # PM file: YAS SID (col 1), HTS SID (col 2), Assigned to (col 3), Planning Date (col 4)
                yas_sid = str(row.iloc[0]) if pd.notna(row.iloc[0]) else None
                helios_sid = str(row.iloc[1]) if pd.notna(row.iloc[1]) else None
                assigned_to = row.iloc[2] if pd.notna(row.iloc[2]) else None
                planning_date = row.iloc[3] if pd.notna(row.iloc[3]) else None
                
                # Convert date to April (normalize all dates to April month)
                if planning_date and hasattr(planning_date, 'month'):
                    # Keep the day/hour info but change month to April
                    april_date = planning_date.replace(month=4, year=2026)
                else:
                    april_date = datetime(2026, 4, 1)
                
                site_id = yas_sid or helios_sid
                if site_id and site_id not in self.high_risk_site_ids and site_id not in processed_sites:
                    processed_sites.add(site_id)
                    
                    # Get site info from Colo_Sonatel lookup
                    site_info = site_master_lookup.get(yas_sid, {}) if yas_sid else {}
                    
                    # Use assigned_to if available, otherwise rotate
                    if assigned_to and pd.notna(assigned_to):
                        tech_staff = str(assigned_to).strip()
                        _, non_tech_1, non_tech_2 = self._assign_staff(visit_count, engineer_pools)
                    else:
                        tech_staff, non_tech_1, non_tech_2 = self._assign_staff(visit_count, engineer_pools)
                    
                    schedule_data.append({
                        'Activity': 'Preventive Maintenance',
                        'Site ID': yas_sid or helios_sid,
                        'Name': site_info.get('name', ''),
                        'Location Type': site_info.get('tower_type', ''),
                        'Latitude': site_info.get('latitude', ''),
                        'Longitude': site_info.get('longitude', ''),
                        'Region': '',
                        'Month': 'April',
                        'Technical Staff': tech_staff,
                        'Non-Tech Staff 1': non_tech_1,
                        'Non-Tech Staff 2': non_tech_2,
                        'Date': april_date,
                        'SARID': ''
                    })
                    visit_count += 1
        
        # ===== PRIORITY 2: Colo_Sonatel (Site Survey activity) =====
        if self.site_master is not None and len(self.site_master) > 0:
            self.log_message(f"  Processing {len(self.site_master)} Site Master entries (Colo Sonatel)...")
            for idx, row in self.site_master.iterrows():
                yas_sid = str(row.iloc[3]) if pd.notna(row.iloc[3]) else None
                
                if yas_sid and yas_sid not in self.high_risk_site_ids and yas_sid not in processed_sites:
                    processed_sites.add(yas_sid)
                    
                    tech_staff, non_tech_1, non_tech_2 = self._assign_staff(visit_count, engineer_pools)
                    
                    # Distribute April dates - roughly 2-3 per week
                    week_num = (visit_count % 4) + 1
                    day_offset = ((visit_count % 3) * 3)
                    april_date = datetime(2026, 4, 1) + timedelta(days=(week_num-1)*7 + day_offset)
                    
                    schedule_data.append({
                        'Activity': 'Site Survey - Colo Sonatel',
                        'Site ID': yas_sid,
                        'Name': row.iloc[4] if pd.notna(row.iloc[4]) else '',
                        'Location Type': row.iloc[2] if pd.notna(row.iloc[2]) else '',
                        'Latitude': '',
                        'Longitude': '',
                        'Region': '',
                        'Month': 'April',
                        'Technical Staff': tech_staff,
                        'Non-Tech Staff 1': non_tech_1,
                        'Non-Tech Staff 2': non_tech_2,
                        'Date': april_date,
                        'SARID': ''
                    })
                    visit_count += 1
        
        # ===== PRIORITY 3: Project Sites BNS plan (with activity column) =====
        if self.project_sites is not None and len(self.project_sites) > 0:
            self.log_message(f"  Processing {len(self.project_sites)} Project Sites (BNS plan)...")
            for idx, row in self.project_sites.iterrows():
                # Find Helios, YAS, Activity, and coordinates
                helios_id = None
                yas_id = None
                lat = None
                lon = None
                activity = 'Project Site'
                
                for col_idx, cell_val in enumerate(row):
                    if pd.isna(cell_val):
                        continue
                    
                    cell_str = str(cell_val).strip()
                    
                    # Activity column (typically has BTS, etc.)
                    if col_idx == 9 and cell_str and cell_str not in ['Activity', 'S/N']:
                        activity = 'Project Site - ' + cell_str
                    
                    # Helios ID pattern (SN...)
                    if cell_str.startswith('SN') and not helios_id:
                        helios_id = cell_str
                    
                    # YAS ID pattern (DK...)
                    elif cell_str.startswith('DK') and not yas_id:
                        yas_id = cell_str
                    
                    # Coordinates
                    elif isinstance(cell_val, (int, float)):
                        try:
                            fval = float(cell_val)
                            if -180 <= fval <= 180:
                                if lat is None and -90 <= fval <= 90:
                                    lat = fval
                                elif lon is None and lat is not None:
                                    lon = fval
                        except:
                            pass
                
                site_id = yas_id or helios_id
                if site_id and site_id not in self.high_risk_site_ids and site_id not in processed_sites:
                    processed_sites.add(site_id)
                    
                    tech_staff, non_tech_1, non_tech_2 = self._assign_staff(visit_count, engineer_pools)
                    
                    # Distribute dates in third week onwards (stay within April 1-30)
                    week_num = ((visit_count % 7) % 2) + 3  # Weeks 3-4
                    day_offset = (visit_count % 7) * 3
                    april_date = datetime(2026, 4, 1) + timedelta(days=(week_num-1)*7 + day_offset)
                    # Clamp to April 1-30
                    if april_date.month != 4:
                        april_date = datetime(2026, 4, 30)
                    
                    schedule_data.append({
                        'Activity': activity,
                        'Site ID': site_id,
                        'Name': '',
                        'Location Type': '',
                        'Latitude': lat or '',
                        'Longitude': lon or '',
                        'Region': '',
                        'Month': 'April',
                        'Technical Staff': tech_staff,
                        'Non-Tech Staff 1': non_tech_1,
                        'Non-Tech Staff 2': non_tech_2,
                        'Date': april_date,
                        'SARID': ''
                    })
                    visit_count += 1
        
        # ===== PRIORITY 4: HTS Calendar (Reference data - safety inspections) =====
        if self.hts_calendar is not None and len(self.hts_calendar) > 0:
            self.log_message(f"  Processing {len(self.hts_calendar)} HTS Calendar entries (Safety Inspections)...")
            for idx, row in self.hts_calendar.iterrows():
                # Extract from HTS Calendar structure
                site_id = None
                activity = ''
                
                for col_idx, cell_val in enumerate(row):
                    if pd.isna(cell_val):
                        continue
                    
                    cell_str = str(cell_val).lower()
                    
                    # Activity column
                    if col_idx == 0 and cell_str and 'preventive' not in cell_str:
                        activity = str(cell_val)
                    
                    # Site ID columns
                    elif str(cell_val).startswith(('DK', 'SN')) and not site_id:
                        site_id = str(cell_val)
                
                if site_id and site_id not in self.high_risk_site_ids and site_id not in processed_sites:
                    processed_sites.add(site_id)
                    
                    tech_staff, non_tech_1, non_tech_2 = self._assign_staff(visit_count, engineer_pools)
                    
                    # Early April dates for HTS
                    april_date = datetime(2026, 4, 1) + timedelta(days=(visit_count % 10))
                    
                    schedule_data.append({
                        'Activity': activity or 'HTS Safety Inspection',
                        'Site ID': site_id,
                        'Name': '',
                        'Location Type': '',
                        'Latitude': '',
                        'Longitude': '',
                        'Region': '',
                        'Month': 'April',
                        'Technical Staff': tech_staff,
                        'Non-Tech Staff 1': non_tech_1,
                        'Non-Tech Staff 2': non_tech_2,
                        'Date': april_date,
                        'SARID': ''
                    })
                    visit_count += 1
        
        self.april_schedule = pd.DataFrame(schedule_data)
        self.log_message(f"✅ April schedule generated: {len(self.april_schedule)} visits (April-only, all sources integrated)")
        return self.april_schedule
    
    def analyze_pm_workload(self) -> pd.DataFrame:
        """Analyze workload by PM zone/assignment."""
        if self.pm_assignments is None:
            return pd.DataFrame()
        
        self.log_message("Analyzing PM workload...")
        
        # Group by assignment column
        assigned_col = [c for c in self.pm_assignments.columns if 'assigned' in c.lower()]
        
        if assigned_col:
            workload = self.pm_assignments.groupby(assigned_col[0]).size().reset_index(name='Count')
            workload['Percentage'] = (workload['Count'] / workload['Count'].sum() * 100).round(2)
            self.log_message(f"✅ PM workload analysis: {len(workload)} assignments")
            return workload
        
        return pd.DataFrame()
    
    def identify_high_priority_sites(self) -> pd.DataFrame:
        """Identify high priority sites requiring immediate attention."""
        high_priority = []
        
        if self.critical_sites is not None:
            # Get sites marked as failed or P0/P1
            priority_col = [c for c in self.critical_sites.columns if 'priority' in c.lower()]
            
            if priority_col:
                critical_sites_list = self.critical_sites[
                    self.critical_sites[priority_col[0]].isin(['P0', 'P1'])
                ]
                high_priority.append(critical_sites_list)
        
        if high_priority:
            result = pd.concat(high_priority, ignore_index=True)
            self.log_message(f"✅ High priority sites identified: {len(result)}")
            return result
        
        return pd.DataFrame()
    
    def create_summary_statistics(self) -> Dict:
        """Create summary statistics for the report."""
        stats = {
            'Report Date': self.report_date.strftime("%B %Y"),
            'Total Sites Scheduled': len(self.april_schedule) if self.april_schedule is not None else 0,
            'High-Risk Sites Excluded': len(self.high_risk_site_ids),
            'Preventive Maintenance': 0,
            'Project Site Visits': 0,
            'Routine Visits': 0,
            'High Priority Count': 0
        }
        
        if self.april_schedule is not None and len(self.april_schedule) > 0:
            # Count by activity type (new data structure)
            stats['Preventive Maintenance'] = (self.april_schedule['Activity'] == 'Preventive Maintenance').sum()
            stats['Project Site Visits'] = (self.april_schedule['Activity'] == 'Project Site Visit').sum()
            stats['Routine Visits'] = (self.april_schedule['Activity'] == 'Routine Visit').sum()
        
        high_priority = self.identify_high_priority_sites()
        stats['High Priority Count'] = len(high_priority)
        
        self.summary_stats = stats
        self.log_message(f"✅ Summary statistics created: {stats}")
        return stats
    
    def create_excel_report(self, output_filename: str = "HT_Site_Visit_Calendar_April_2026.xlsx"):
        """Create comprehensive Excel report with multiple sheets."""
        self.log_message(f"Creating Excel report: {output_filename}")
        
        output_path = self.output_folder / output_filename
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Summary", 0)
        self._write_summary_sheet(ws_summary)
        
        # 2. April Schedule Sheet
        if self.april_schedule is not None:
            ws_schedule = wb.create_sheet("April Schedule", 1)
            self._write_dataframe_sheet(ws_schedule, self.april_schedule, "Site Visit Schedule")
        
        # 3. High Priority Sites
        high_priority = self.identify_high_priority_sites()
        if not high_priority.empty:
            ws_priority = wb.create_sheet("High Priority", 2)
            self._write_dataframe_sheet(ws_priority, high_priority.head(50), "High Priority Sites")
        
        # 4. PM Workload Analysis
        pm_workload = self.analyze_pm_workload()
        if not pm_workload.empty:
            ws_pm = wb.create_sheet("PM Workload", 3)
            self._write_dataframe_sheet(ws_pm, pm_workload, "PM Assignment Summary")
        
        # 5. Analysis Log
        ws_log = wb.create_sheet("Analysis Log", 4)
        self._write_log_sheet(ws_log)
        
        # Save
        wb.save(output_path)
        self.log_message(f"✅ Report saved: {output_path}")
        return output_path
    
    def _write_summary_sheet(self, ws):
        """Write summary statistics to sheet."""
        ws['A1'] = 'HTS SITE VISIT CALENDAR - APRIL 2026'
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = 'Report Summary'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 2
        if self.summary_stats:
            for key, value in self.summary_stats.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _write_dataframe_sheet(self, ws, df: pd.DataFrame, title: str):
        """Write dataframe to sheet with formatting."""
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A1:{chr(65 + len(df.columns) - 1)}1')
        
        # Write dataframe
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format header
                if r_idx == 2:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Auto-fit columns
        for col_idx, col in enumerate(df.columns, 1):
            try:
                col_values = df[col].fillna('').astype(str)
                max_len = max(col_values.map(len).max() if col_values.map(len).max() > 0 else 0, len(col))
            except:
                max_len = len(col)
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 2, 40)
    
    def _write_log_sheet(self, ws):
        """Write analysis log to sheet."""
        ws['A1'] = 'Analysis Log'
        ws['A1'].font = Font(size=12, bold=True)
        
        for idx, msg in enumerate(self.log, 2):
            ws[f'A{idx}'] = msg
        
        ws.column_dimensions['A'].width = 100
    
    def run_analysis(self, files: Dict[str, str], output_filename: str = "HT_Site_Visit_Calendar_April_2026.xlsx") -> bool:
        """Execute complete analysis pipeline."""
        self.log_message("Starting April Site Visit Analysis...")
        
        if not self.load_files(files):
            return False
        
        self.generate_april_schedule()
        self.create_summary_statistics()
        self.create_excel_report(output_filename)
        
        self.log_message("✅ Analysis complete!")
        return True


def main():
    parser = argparse.ArgumentParser(
        description='Generate April Site Visit Calendar from 5 file types'
    )
    parser.add_argument('--site-master', type=str, help='Path to site master file')
    parser.add_argument('--hts-calendar', type=str, help='Path to HTS calendar file')
    parser.add_argument('--pm-assignments', type=str, help='Path to PM assignments file')
    parser.add_argument('--project-sites', type=str, help='Path to project sites file')
    parser.add_argument('--critical-sites', type=str, help='Path to critical sites file')
    parser.add_argument('--output', type=str, default='output', help='Output folder')
    parser.add_argument('--report-name', type=str, default='HT_Site_Visit_Calendar_April_2026.xlsx',
                       help='Output report filename')
    
    args = parser.parse_args()
    
    # Create file dictionary with defaults
    files = {}
    
    # Set default paths based on typical structure
    base_path = Path(__file__).parent.parent / 'data'
    
    files['site_master'] = args.site_master or str(base_path / 'Colo_Sonatel-APS.xlsx')
    files['hts_calendar'] = args.hts_calendar or str(base_path / 'HTS Site Visits Calendar March 26 R0 (1).xlsx')
    files['pm_assignments'] = args.pm_assignments or str(base_path / 'PM_ZONE1_2_MARS_2026.xlsx')
    files['project_sites'] = args.project_sites or str(base_path / 'Project sites BNS plan for March 2026.xlsx')
    files['critical_sites'] = args.critical_sites or str(base_path / 'RT High Risk List_of_sites_BE_failed 1.xlsx')
    
    # Run generator
    generator = AprilReportGenerator(output_folder=args.output)
    success = generator.run_analysis(files, args.report_name)
    
    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
