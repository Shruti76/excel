#!/usr/bin/env python3
"""
Balanced Engineer Assignment Script
Redistributes sites equally among all 40 engineers based on daily and weekly schedules
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random


class BalancedEngineersScheduler:
    """Create balanced engineer assignments by day and week"""
    
    def __init__(self, report_path: str):
        self.report_path = Path(report_path)
        self.schedule_df = None
        self.engineers = None
        self.output_folder = Path('/Users/shrutisohan/Desktop/excel/output')
        
    def load_data(self):
        """Load schedule and engineer list from report"""
        # Load PM Workload to get engineers
        pm_workload = pd.read_excel(self.report_path, sheet_name='PM Workload', header=None)
        
        # Extract engineer names (skip header rows)
        self.engineers = pm_workload.iloc[1:, 0].tolist()
        self.engineers = [e for e in self.engineers if e and str(e).strip()]
        
        print(f"✅ Loaded {len(self.engineers)} engineers")
        print(f"Engineers: {self.engineers[:5]}...")
    
    def create_balanced_schedule(self, input_file: str) -> pd.DataFrame:
        """
        Create balanced daily schedule
        Assigns engineers evenly across all 4 weeks with balanced daily loads
        """
        
        # Load original schedule
        schedule_df = pd.read_excel(input_file, sheet_name='April Schedule')
        
        # Extract actual columns
        schedule_df.columns = schedule_df.iloc[0]
        schedule_df = schedule_df[1:].reset_index(drop=True)
        
        print(f"\n📅 Schedule Analysis:")
        print(f"   Total sites: {len(schedule_df)}")
        
        # Parse scheduled dates and create daily assignments
        assignments = []
        engineers_cycle = self.engineers.copy()
        random.shuffle(engineers_cycle)
        engineer_idx = 0
        
        # Group by week
        weeks = {}
        for idx, row in schedule_df.iterrows():
            week = row.get('Scheduled Week', 'Week 1')
            if week not in weeks:
                weeks[week] = []
            weeks[week].append(idx)
        
        print(f"   Weeks: {list(weeks.keys())}")
        
        # Assign engineers round-robin by week
        engineer_assignments = []
        for week in sorted(weeks.keys()):
            week_sites = weeks[week]
            print(f"\n   {week}: {len(week_sites)} sites")
            
            # Distribute engineers evenly within week
            for site_idx in week_sites:
                engineer = engineers_cycle[engineer_idx % len(engineers_cycle)]
                engineer_assignments.append({
                    'site_idx': site_idx,
                    'engineer': engineer,
                    'week': week
                })
                engineer_idx += 1
        
        # Apply assignments to dataframe
        assignment_dict = {a['site_idx']: a['engineer'] for a in engineer_assignments}
        schedule_df['Engineer'] = schedule_df.index.map(lambda x: assignment_dict.get(x, ''))
        
        return schedule_df
    
    def analyze_daily_distribution(self, schedule_df: pd.DataFrame):
        """Analyze how engineers are distributed by day"""
        
        print("\n" + "=" * 100)
        print("DAILY DISTRIBUTION ANALYSIS")
        print("=" * 100)
        
        # Group by week and day within week
        daily_counts = {}
        weekly_counts = {}
        
        for week in ['Week 1', 'Week 2', 'Week 3', 'Week 4']:
            week_data = schedule_df[schedule_df['Scheduled Week'] == week]
            weekly_counts[week] = len(week_data)
            
            print(f"\n{week} ({len(week_data)} total sites):")
            
            # Distribute within week
            days_in_week = 5  # Mon-Fri
            sites_per_day = len(week_data) // days_in_week
            remainder = len(week_data) % days_in_week
            
            daily_distribution = []
            for day in range(days_in_week):
                day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'][day]
                sites = sites_per_day + (1 if day < remainder else 0)
                daily_distribution.append({
                    'day': day_name,
                    'sites': sites
                })
                print(f"   {day_name}: {sites} sites")
        
        return weekly_counts, daily_distribution
    
    def create_engineer_daily_schedule(self, schedule_df: pd.DataFrame) -> pd.DataFrame:
        """Create detailed engineer schedule by day"""
        
        print("\n" + "=" * 100)
        print("ENGINEER DAILY ASSIGNMENT SCHEDULE")
        print("=" * 100)
        
        daily_schedule = []
        
        for week in ['Week 1', 'Week 2', 'Week 3', 'Week 4']:
            week_data = schedule_df[schedule_df['Scheduled Week'] == week].copy()
            
            # Create daily distribution
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            sites_per_day = len(week_data) // 5
            remainder = len(week_data) % 5
            
            site_idx = 0
            
            for day_num, day_name in enumerate(days):
                # Calculate sites for this day
                day_sites = sites_per_day + (1 if day_num < remainder else 0)
                
                # Get sites for this day
                day_assignments = []
                for i in range(day_sites):
                    if site_idx < len(week_data):
                        row = week_data.iloc[site_idx]
                        engineer = row.get('Engineer', '')
                        
                        daily_schedule.append({
                            'Week': week,
                            'Day': day_name,
                            'Site ID': row.get('Site ID', ''),
                            'Engineer': engineer,
                            'Visit Type': row.get('Visit Type', ''),
                            'Priority': row.get('Priority', ''),
                            'Duration (hours)': row.get('Duration (hours)', 2)
                        })
                        site_idx += 1
        
        daily_df = pd.DataFrame(daily_schedule)
        
        print(f"\n📋 Daily Schedule Created: {len(daily_df)} assignments")
        
        return daily_df
    
    def get_engineer_balance_report(self, schedule_df: pd.DataFrame) -> pd.DataFrame:
        """Generate engineer workload balance report"""
        
        print("\n" + "=" * 100)
        print("ENGINEER WORKLOAD BALANCE REPORT")
        print("=" * 100)
        
        # Count assignments per engineer
        engineer_counts = schedule_df['Engineer'].value_counts().reset_index()
        engineer_counts.columns = ['Engineer', 'Sites Assigned']
        engineer_counts['Percentage'] = (
            engineer_counts['Sites Assigned'] / engineer_counts['Sites Assigned'].sum() * 100
        ).round(2)
        
        # Calculate statistics
        mean_sites = engineer_counts['Sites Assigned'].mean()
        std_sites = engineer_counts['Sites Assigned'].std()
        max_sites = engineer_counts['Sites Assigned'].max()
        min_sites = engineer_counts['Sites Assigned'].min()
        
        print(f"\n📊 Workload Statistics:")
        print(f"   Average sites per engineer: {mean_sites:.2f}")
        print(f"   Standard deviation: {std_sites:.2f}")
        print(f"   Max assignments: {max_sites}")
        print(f"   Min assignments: {min_sites}")
        print(f"   Balance ratio (Max/Min): {max_sites/min_sites:.2f}")
        
        print(f"\n👥 Top 10 Engineers by Assignment:")
        print(engineer_counts.head(10).to_string(index=False))
        
        print(f"\n👥 Bottom 10 Engineers by Assignment:")
        print(engineer_counts.tail(10).to_string(index=False))
        
        return engineer_counts
    
    def create_excel_output(self, schedule_df: pd.DataFrame, daily_df: pd.DataFrame, 
                           balance_df: pd.DataFrame):
        """Create comprehensive Excel output with all schedules"""
        
        output_path = self.output_folder / 'Engineer_Assignment_Schedule_April_2026.xlsx'
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Balanced Schedule
            schedule_df.to_excel(writer, sheet_name='Balanced Schedule', index=False)
            
            # Sheet 2: Daily Schedule
            daily_df.to_excel(writer, sheet_name='Daily Schedule', index=False)
            
            # Sheet 3: Engineer Workload
            balance_df.to_excel(writer, sheet_name='Engineer Workload', index=False)
            
            # Sheet 4: Summary
            summary_data = {
                'Metric': [
                    'Total Sites',
                    'Total Engineers',
                    'Avg Sites per Engineer',
                    'Max Assignment',
                    'Min Assignment',
                    'Balance Ratio'
                ],
                'Value': [
                    len(schedule_df),
                    len(balance_df),
                    f"{schedule_df['Engineer'].value_counts().mean():.2f}",
                    schedule_df['Engineer'].value_counts().max(),
                    schedule_df['Engineer'].value_counts().min(),
                    f"{schedule_df['Engineer'].value_counts().max() / schedule_df['Engineer'].value_counts().min():.2f}"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format the Excel file
        wb = load_workbook(output_path)
        
        for sheet in ['Balanced Schedule', 'Daily Schedule', 'Engineer Workload', 'Summary']:
            ws = wb[sheet]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        wb.save(output_path)
        print(f"\n✅ Excel file saved: {output_path}")
        
        return output_path


def main():
    report_path = '/Users/shrutisohan/Desktop/excel/output/HT_Site_Visit_Calendar_April_2026.xlsx'
    original_schedule = '/Users/shrutisohan/Desktop/excel/output/HT_Site_Visit_Calendar_April_2026.xlsx'
    
    scheduler = BalancedEngineersScheduler(report_path)
    scheduler.load_data()
    
    # Create balanced schedule
    balanced_schedule = scheduler.create_balanced_schedule(original_schedule)
    
    # Analyze distribution
    scheduler.analyze_daily_distribution(balanced_schedule)
    
    # Create daily schedule
    daily_schedule = scheduler.create_engineer_daily_schedule(balanced_schedule)
    
    # Get balance report
    balance_report = scheduler.get_engineer_balance_report(balanced_schedule)
    
    # Create Excel output
    scheduler.create_excel_output(balanced_schedule, daily_schedule, balance_report)
    
    print("\n" + "=" * 100)
    print("✅ BALANCED ENGINEER SCHEDULE CREATED SUCCESSFULLY")
    print("=" * 100)


if __name__ == '__main__':
    main()
